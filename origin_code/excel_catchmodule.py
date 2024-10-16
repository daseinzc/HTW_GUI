from openpyxl import load_workbook

def excelcatch():
    wb = load_workbook('clickme.xlsx')
    ws = wb.active
    ls_hang = []
    ls_school = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ls_hang.append(row)
        school,year,month,money,due_day = row
        ls_school.append(school)
    return ls_hang,ls_school

