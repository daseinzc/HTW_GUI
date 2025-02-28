import sys
import os
import json
import logging
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout,
    QPushButton, QWidget, QHBoxLayout, QAbstractItemView, QFileDialog, QMessageBox,
    QLineEdit, QStackedWidget, QLabel, QHeaderView, QAction, QFrame, QMenu, QStyle
)
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QIcon, QFont, QKeySequence
from qt_material import apply_stylesheet
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from plugin_page import PluginPage  # 导入 PluginPage 类
import ctypes

# 设置明确的Windows应用ID (这会强制Windows使用新图标)
try:
    # 注意：这个ID必须唯一且保持一致，不要随意更改
    app_id = u'HUST.TeamFeeSystem.HGUI.2025'
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
except Exception as e:
    print(f"设置应用ID出错: {e}")

# 添加资源路径解析函数
def get_resource_path(relative_path):
    """获取资源的绝对路径，适用于开发环境和PyInstaller打包环境"""
    # PyInstaller会创建临时文件夹并将路径存储在_MEIPASS中
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)


# 配置日志记录（可选）
os.makedirs('logs', exist_ok=True)
logging.basicConfig(
    filename='logs/run.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)



class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            self_val = float(self.text())
        except ValueError:
            self_val = 0.0
        try:
            other_val = float(other.text())
        except ValueError:
            other_val = 0.0
        return self_val < other_val

class ModernTableApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # 初始化撤回历史记录（提前初始化）
        self.undo_stack = []
        self.max_undo_steps = 60  # 最多保存60步操作

        # 1) 全局字体
        font = QFont("Roboto", 12)
        font.setStyleHint(QFont.SansSerif)
        font.setWeight(QFont.Normal)
        app.setFont(font)

        self.setWindowTitle("基团团费整理系统")
        self.setGeometry(100, 100, 1400, 800)
        self.setWindowFlags(Qt.Window | Qt.WindowTitleHint | Qt.WindowMinMaxButtonsHint | Qt.WindowCloseButtonHint)

        # 设置图标（可选）
        icon_path = get_resource_path(os.path.join("image", "pic.ico"))
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        # 2) 中央布局：左侧导航(窄) + 右侧StackedWidget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0,0,0,0)

        # 左侧导航栏(更窄)
        self.nav_widget = QWidget()
        self.nav_widget.setFixedWidth(60)  # 仅容纳图标
        self.nav_widget.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5; 
                border-right: 1px solid #e0e0e0;
            }
        """)
        self.nav_layout = QVBoxLayout(self.nav_widget)
        self.nav_layout.setContentsMargins(0,10,0,10)
        self.nav_layout.setSpacing(10)

        # 右侧 StackedWidget
        self.stacked_widget = QStackedWidget()
        self.main_page = QWidget(self)
        self.option_page = QWidget(self)
        self.plugin_page = PluginPage(main_page=self.main_page, stacked_widget=self.stacked_widget)

        self.stacked_widget.addWidget(self.main_page)
        self.stacked_widget.addWidget(self.option_page)
        self.stacked_widget.addWidget(self.plugin_page)
        self.stacked_widget.setCurrentWidget(self.main_page)

        main_layout.addWidget(self.nav_widget)
        main_layout.addWidget(self.stacked_widget)

        # ========== 导航栏按钮(纯图标) ==========
        # 3) 加载图标(黑色线条),你可用 home_black.svg/settings_black.svg/plugin_black.svg
        # 加载图标，使用异常处理
        try:
            # 3) 加载图标(黑色线条)，使用resource_path确保在打包环境中可用
            icon_home = QIcon(get_resource_path(os.path.join("image", "home.svg")))
            icon_settings = QIcon(get_resource_path(os.path.join("image", "settings.svg")))
            icon_plugin = QIcon(get_resource_path(os.path.join("image", "plugin.svg")))

            # 记录日志
            logging.debug(f"导航图标加载成功")
        except Exception as e:
            logging.error(f"加载导航图标时发生错误: {e}")
            # 使用空图标作为后备
            icon_home = QIcon()
            icon_settings = QIcon()
            icon_plugin = QIcon()

        self.btn_go_main = QPushButton()
        self.btn_go_main.setIcon(icon_home)
        self.btn_go_main.setIconSize(QSize(24,24))
        self.btn_go_main.setToolTip("首页")

        self.btn_go_option = QPushButton()
        self.btn_go_option.setIcon(icon_settings)
        self.btn_go_option.setIconSize(QSize(24,24))
        self.btn_go_option.setToolTip("选项页")

        self.btn_go_plugin = QPushButton()
        self.btn_go_plugin.setIcon(icon_plugin)
        self.btn_go_plugin.setIconSize(QSize(24,24))
        self.btn_go_plugin.setToolTip("插件管理")

        # 4) QSS: 透明默认背景 + 悬浮/按下显示灰色, 让图标有焦点反馈
        #    并保持按钮区域小, padding: 8px, 这样鼠标点击范围大些
        btn_style = """
            QPushButton {
                background-color: transparent; 
                border: none;
                padding: 8px; 
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
            QPushButton:pressed {
                background-color: #cccccc;
            }
            QPushButton:focus {
                outline: none;
                border: none;
            }
        """
        for btn in [self.btn_go_main, self.btn_go_option, self.btn_go_plugin]:
            btn.setStyleSheet(btn_style)

        # 5) 点击事件切换页面
        self.btn_go_main.clicked.connect(lambda: self.stacked_widget.setCurrentWidget(self.main_page))
        self.btn_go_option.clicked.connect(lambda: self.stacked_widget.setCurrentWidget(self.option_page))
        self.btn_go_plugin.clicked.connect(lambda: self.stacked_widget.setCurrentWidget(self.plugin_page))

        # 加入左侧布局
        self.nav_layout.addWidget(self.btn_go_main)
        self.nav_layout.addWidget(self.btn_go_option)
        self.nav_layout.addWidget(self.btn_go_plugin)
        self.nav_layout.addStretch(1)

        # 初始化三个页面 + 菜单
        self.init_main_page()
        self.init_option_page()
        self.init_menu()

        # 创建状态栏
        self.statusBar().showMessage("就绪")
        self.statusBar().setStyleSheet("font-size: 14px; padding: 4px;")

        self._is_undoing = False
    def init_menu(self):
        menubar = self.menuBar()
        menubar.setStyleSheet("font-size: 16px; background-color: #ffffff; color: #333333;")

        # 文件菜单
        file_menu = menubar.addMenu("文件")

        save_action = QAction("保存进度", self)
        try:
            save_menu_icon_path = get_resource_path(os.path.join("image", "save.svg"))
            if os.path.exists(save_menu_icon_path):
                save_action.setIcon(QIcon(save_menu_icon_path))
            else:
                save_action.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogSaveButton))
        except Exception as e:
            logging.error(f"加载菜单保存图标时发生错误: {e}")
            save_action.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogSaveButton))
        save_action.triggered.connect(self.save_progress)
        save_action.setShortcut(QKeySequence.Save)  # Ctrl+S
        file_menu.addAction(save_action)

        load_action = QAction("导入进度", self)
        load_action.setShortcut(QKeySequence("Ctrl+L"))
        load_action.triggered.connect(self.load_progress)
        file_menu.addAction(load_action)

        file_menu.addSeparator()

        excel_action = QAction("输出为Excel", self)
        excel_action.setShortcut(QKeySequence("Ctrl+E"))
        excel_action.triggered.connect(self.output_to_excel)
        file_menu.addAction(excel_action)

        # 添加帮助菜单
        help_menu = menubar.addMenu("帮助")

        user_guide_action = QAction("用户指南", self)
        user_guide_action.setShortcut(QKeySequence("F1"))
        user_guide_action.triggered.connect(self.show_user_guide)
        help_menu.addAction(user_guide_action)

        help_menu.addSeparator()

        about_action = QAction("关于软件", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def show_user_guide(self):
        """显示用户指南（移除问号按钮）"""
        guide_text = """
    <h2>基团团费整理系统使用指南</h2>

    <h3>基本操作</h3>
    <ul>
      <li><b>添加行</b>：选中某行后点击"添加行"按钮在该行后添加新行</li>
      <li><b>删除行</b>：选中行后点击"删除行"按钮删除整行</li>
      <li><b>编辑单元格</b>：双击单元格或选中后按 F2 开始编辑</li>
      <li><b>选择多个单元格</b>：按住 Ctrl 键点选多个单元格，或拖动鼠标框选</li>
      <li><b>删除单元格内容</b>：选中单元格后按 Delete 键清空内容</li>
    </ul>

    <h3>快捷键</h3>
    <table border="1" cellpadding="4" style="border-collapse: collapse;">
      <tr><td><b>Ctrl+C</b></td><td>复制选中内容</td></tr>
      <tr><td><b>Ctrl+V</b></td><td>粘贴内容</td></tr>
      <tr><td><b>Ctrl+X</b></td><td>剪切选中内容</td></tr>
      <tr><td><b>Ctrl+Z</b></td><td>撤销上一步操作</td></tr>
      <tr><td><b>Ctrl+S</b></td><td>保存进度</td></tr>
      <tr><td><b>Ctrl+L</b></td><td>加载已保存进度</td></tr>
      <tr><td><b>Ctrl+E</b></td><td>输出为Excel</td></tr>
      <tr><td><b>Ctrl+A</b></td><td>全选单元格</td></tr>
      <tr><td><b>Ctrl+O</b></td><td>打开Excel文件</td></tr>
      <tr><td><b>F1</b></td><td>显示用户指南</td></tr>
      <tr><td><b>Delete</b></td><td>删除选中单元格内容</td></tr>
      <tr><td><b>Tab</b></td><td>移动到下一个单元格</td></tr>
      <tr><td><b>Enter</b></td><td>移动到下一行</td></tr>
    </table>

    <h3>文件操作</h3>
    <ul>
      <li><b>保存进度</b>：将当前编辑状态保存为JSON文件，可以之后继续编辑</li>
      <li><b>加载进度</b>：读取之前保存的JSON文件，恢复编辑状态</li>
      <li><b>打开Excel</b>：导入Excel文件数据到表格中</li>
      <li><b>输出为Excel</b>：将当前表格数据导出为Excel文件</li>
    </ul>

    <h3>插件功能</h3>
    <p>通过插件管理页面，您可以：</p>
    <ul>
      <li><b>添加插件</b>：点击"添加插件"按钮，选择外部可执行文件(.exe)作为插件</li>
      <li><b>运行插件</b>：点击插件对应的"运行"按钮启动选定的插件程序</li>
      <li><b>删除插件</b>：点击插件对应的"删除"按钮从列表中移除插件</li>
    </ul>
    <p>插件功能允许您扩展系统功能，例如添加数据分析、报表生成等工具。</p>

    <h3>其他功能</h3>
    <ul>
      <li><b>排序</b>：点击"按序号排序"按钮按序号列对表格进行排序</li>
      <li><b>撤销</b>：点击"撤回操作"按钮或按Ctrl+Z撤销上一步操作</li>
      <li><b>选项设置</b>：在选项页面设置团费年份、月份和落款日期</li>
    </ul>
    """

        # 创建自定义窗口显示格式化的HTML内容
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTextBrowser

        html_dialog = QDialog(self)
        # 移除问号按钮
        html_dialog.setWindowFlags(html_dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        html_dialog.setWindowTitle("基团团费整理系统 - 用户指南")
        html_dialog.resize(700, 600)

        layout = QVBoxLayout(html_dialog)
        text_browser = QTextBrowser()
        text_browser.setHtml(guide_text)
        text_browser.setOpenExternalLinks(True)
        layout.addWidget(text_browser)

        html_dialog.exec_()

    def show_about(self):
        """显示关于软件信息（移除问号按钮）"""
        about_text = f"""
        <div style="text-align: center;">
            <h2>基团团费整理系统</h2>
            <p>版本: 2.0.0</p>
            <p>一个简单高效的团费管理与整理工具</p>
            <p>基于PyQt5开发</p>
            <p>MIT协议</p>
        </div>
        """

        # 创建自定义窗口显示格式化的HTML内容
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTextBrowser

        html_dialog = QDialog(self)
        # 移除问号按钮
        html_dialog.setWindowFlags(html_dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        html_dialog.setWindowTitle("关于基团团费整理系统")
        html_dialog.resize(400, 300)

        layout = QVBoxLayout(html_dialog)
        text_browser = QTextBrowser()
        text_browser.setHtml(about_text)
        layout.addWidget(text_browser)

        html_dialog.exec_()

    def init_main_page(self):
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["序号", "学院", "财务金额", "是否补交"])

        # 增强表格设置 - Excel风格功能
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)  # 允许选择单个单元格
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 允许多选
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)  # 启用右键菜单
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.horizontalHeader().setSortIndicatorShown(False)  # 显示排序指示器

        # 连接单元格编辑信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 优化表格样式
        self.table.setStyleSheet("""
            QTableWidget {
                font-size: 24px; 
                selection-background-color: #a5d6a7;
                gridline-color: #d0d0d0;
                border: 1px solid #c0c0c0;
            }
            QHeaderView::section {
                background-color: #e8f5e9;
                padding: 6px;
                font-weight: bold;
                border: 1px solid #c0c0c0;
            }
            QTableWidget::item {
                padding: 4px;
            }
        """)

        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 设置键盘快捷键
        self.setup_keyboard_shortcuts()

        self.table.selectionModel().selectionChanged.connect(self.update_buttons_state)
        # 填充学院数据
        self.schools = [
            "社会学院", "网络空间安全学院", "光学与电子信息学院", "人文学院", "未来技术学院", "机械科学与工程学院",
            "土木与水利工程学院", "马克思主义学院", "口腔医学院", "能源与动力工程学院", "外国语学院", "护理学院",
            "化学与化工学院", "新闻与信息传播学院", "生命科学与技术学院", "电子信息与通信学院", "法学院", "哲学学院",
            "航空航天学院", "人工智能与自动化学院", "医药卫生管理学院", "物理学院", "管理学院", "环境科学与工程学院",
            "公共管理学院", "第一临床学院", "计算机科学与技术学院", "艺术学院", "法医学系", "数学与统计学院",
            "第二临床学院", "电气与电子工程学院", "经济学院", "集成电路学院", "药学院", "公共卫生学院", "基础医学院",
            "材料科学与工程学院", "建筑与城市规划学院", "船舶与海洋工程学院", "武汉光电国家研究中心", "体育学院",
            "教育科学研究院", "生殖健康研究所", "软件学院"
        ]
        for i, school in enumerate(self.schools):
            self.table.insertRow(i)
            seq_item = NumericTableWidgetItem(str(i + 1))
            self.table.setItem(i, 0, seq_item)
            self.table.setItem(i, 1, QTableWidgetItem(school))
            self.table.setItem(i, 2, QTableWidgetItem(""))
            self.table.setItem(i, 3, QTableWidgetItem(""))

        # 创建底部按钮组，使用图标和更现代的设计
        base_path = os.path.dirname(os.path.abspath(__file__))

        # "添加行"按钮
        self.add_row_btn = QPushButton("添加行")
        try:
            add_row_icon_path = get_resource_path(os.path.join("image", "add_row.svg"))
            if os.path.exists(add_row_icon_path):
                self.add_row_btn.setIcon(QIcon(add_row_icon_path))
                self.add_row_btn.setIconSize(QSize(20, 20))
            else:
                self.add_row_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        except Exception as e:
            logging.error(f"加载添加行图标时发生错误: {e}")
            self.add_row_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogNewFolder))

        self.add_row_btn.setStyleSheet("""
            QPushButton {
                background-color: #4caf50; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #4caf50ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #4caf50cc;  /* 使用透明度创建按下效果 */
            }
        """)

        # "删除行"按钮
        self.del_row_btn = QPushButton("删除行")
        try:
            delete_row_icon_path = get_resource_path(os.path.join("image", "delete_row.svg"))
            if os.path.exists(delete_row_icon_path):
                self.del_row_btn.setIcon(QIcon(delete_row_icon_path))
                self.del_row_btn.setIconSize(QSize(20, 20))
            else:
                self.del_row_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_TrashIcon))
        except Exception as e:
            logging.error(f"加载删除行图标时发生错误: {e}")
            self.del_row_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_TrashIcon))

        self.del_row_btn.setStyleSheet("""
            QPushButton {
                background-color: #e57373; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #e57373ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #e57373cc;  /* 使用透明度创建按下效果 */
            }
        """)

        # "按序号排序"按钮
        self.sort_btn = QPushButton("按序号排序")

        try:
            sort_icon_path = get_resource_path(os.path.join("image", "sort.svg"))
            if os.path.exists(sort_icon_path):
                self.sort_btn.setIcon(QIcon(sort_icon_path))
                self.sort_btn.setIconSize(QSize(20, 20))
            else:
                self.sort_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowDown))
        except Exception as e:
            logging.error(f"加载排序图标时发生错误: {e}")
            self.sort_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_ArrowDown))
        self.sort_btn.setStyleSheet("""
            QPushButton {
                background-color: #64b5f6; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #64b5f6ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #64b5f6cc;  /* 使用透明度创建按下效果 */
            }
        """)

        # 创建带下拉菜单的文件按钮
        self.file_btn = QPushButton("打开文件")
        try:
            open_file_icon_path = get_resource_path(os.path.join("image", "open_file.svg"))
            if os.path.exists(open_file_icon_path):
                self.file_btn.setIcon(QIcon(open_file_icon_path))
                self.file_btn.setIconSize(QSize(20, 20))
            else:
                self.file_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogOpenButton))
        except Exception as e:
            logging.error(f"加载打开文件图标时发生错误: {e}")
            self.file_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogOpenButton))

        self.file_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff9800; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
                text-align: left;
                padding-right: 15px;
            }
            QPushButton:hover {
                background-color: #ff9800ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #ff9800cc;  /* 使用透明度创建按下效果 */
            }
            QPushButton::menu-indicator {
                subcontrol-position: right center;
                subcontrol-origin: padding;
                right: 8px;
            }
        """)

        # 创建下拉菜单
        file_menu = QMenu(self)
        file_menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #cccccc;
                padding: 5px;
                border-radius: 3px;
            }
            QMenu::item {
                padding: 6px 25px 6px 20px;
                border-radius: 3px;
            }
            QMenu::item:selected {
                background-color: #f0f0f0;
            }
        """)

        open_excel_action = file_menu.addAction("打开Excel文件")
        try:
            open_excel_icon_path = get_resource_path(os.path.join("image", "excel_dark.svg"))
            if os.path.exists(open_excel_icon_path):
                open_excel_action.setIcon(QIcon(open_excel_icon_path))
            else:
                open_excel_action.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        except Exception as e:
            logging.error(f"加载打开Excel菜单图标时发生错误: {e}")
            open_excel_action.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogDetailedView))

        load_progress_action = file_menu.addAction("加载已保存进度")
        try:
            load_icon_path = get_resource_path(os.path.join("image", "load.svg"))
            if os.path.exists(load_icon_path):
                load_progress_action.setIcon(QIcon(load_icon_path))
            else:
                load_progress_action.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogContentsView))
        except Exception as e:
            logging.error(f"加载进度加载图标时发生错误: {e}")
            load_progress_action.setIcon(QApplication.style().standardIcon(QStyle.SP_FileDialogContentsView))

        self.file_btn.setMenu(file_menu)

        # 创建保存进度按钮
        self.save_btn = QPushButton("保存进度")
        # 详细的图标和路径调试
        logging.debug("调试：保存按钮图标加载")
        base_path = os.path.dirname(os.path.abspath(__file__))
        logging.debug(f"当前脚本路径: {base_path}")

        try:
            # 使用新的resource path获取方法
            save_icon_path = get_resource_path(os.path.join("image", "save.svg"))
            logging.debug(f"尝试加载图标路径: {save_icon_path}")
            logging.debug(f"文件是否存在: {os.path.exists(save_icon_path)}")

            if os.path.exists(save_icon_path):
                icon = QIcon(save_icon_path)
                if not icon.isNull():
                    self.save_btn.setIcon(icon)
                    self.save_btn.setIconSize(QSize(20, 20))
                    logging.debug("成功加载自定义图标")
                else:
                    logging.warning("警告：图标加载失败（QIcon为空）")
                    self.save_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogSaveButton))
            else:
                logging.warning(f"未找到图标文件: {save_icon_path}")
                self.save_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogSaveButton))
        except Exception as e:
            logging.error(f"加载图标时发生错误: {e}")
            self.save_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_DialogSaveButton))

        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #9c27b0; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #9c27b0ee;
            }
            QPushButton:pressed {
                background-color: #9c27b0cc;
            }
        """)

        # 创建输出Excel按钮
        self.output_btn = QPushButton("输出为Excel")

        try:
            excel_icon_path = get_resource_path(os.path.join("image", "excel.svg"))
            if os.path.exists(excel_icon_path):
                self.output_btn.setIcon(QIcon(excel_icon_path))
                self.output_btn.setIconSize(QSize(20, 20))
            else:
                self.output_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileIcon))
        except Exception as e:
            logging.error(f"加载Excel图标时发生错误: {e}")
            self.output_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_FileIcon))

        self.output_btn.setStyleSheet("""
            QPushButton {
                background-color: #81c784; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #81c784ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #81c784cc;  /* 使用透明度创建按下效果 */
            }
        """)

        # 撤回按钮
        self.undo_btn = QPushButton("撤回操作")

        try:
            undo_icon_path = get_resource_path(os.path.join("image", "undo.svg"))
            if os.path.exists(undo_icon_path):
                self.undo_btn.setIcon(QIcon(undo_icon_path))
                self.undo_btn.setIconSize(QSize(20, 20))
            else:
                self.undo_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        except Exception as e:
            logging.error(f"加载撤回图标时发生错误: {e}")
            self.undo_btn.setIcon(QApplication.style().standardIcon(QStyle.SP_BrowserReload))
        self.undo_btn.setStyleSheet("""
            QPushButton {
                background-color: #ba68c8; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #ba68c8ee;  /* 使用透明度创建悬停效果 */
            }
            QPushButton:pressed {
                background-color: #ba68c8cc;  /* 使用透明度创建按下效果 */
            }
        """)

        # 初始时禁用添加和删除按钮，直到有选择
        self.add_row_btn.setEnabled(False)
        self.del_row_btn.setEnabled(False)

        # 改进按钮布局，使用更现代的分组设计
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)  # 按钮之间的间距

        # 按功能分组的按钮
        edit_group = QHBoxLayout()
        edit_group.addWidget(self.add_row_btn)
        edit_group.addWidget(self.del_row_btn)
        edit_group.addWidget(self.undo_btn)

        file_group = QHBoxLayout()
        file_group.addWidget(self.sort_btn)
        file_group.addWidget(self.file_btn)  # 新的文件下拉按钮
        file_group.addWidget(self.save_btn)  # 新的保存进度按钮
        file_group.addWidget(self.output_btn)

        # 向主布局添加分组
        btn_layout.addLayout(edit_group)
        btn_layout.addStretch(1)  # 弹性空间
        btn_layout.addLayout(file_group)

        # 主页面布局
        main_layout = QVBoxLayout(self.main_page)
        header_label = QLabel("基团团费整理系统", self)
        header_label.setStyleSheet("font-size: 28px; font-weight: bold; padding: 15px; color: #2e7d32;")
        header_label.setAlignment(Qt.AlignCenter)

        # 添加分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("background-color: #c8e6c9; margin: 0 15px;")

        main_layout.addWidget(header_label)
        main_layout.addWidget(line)
        main_layout.addWidget(self.table)
        main_layout.addLayout(btn_layout)

        # 信号连接
        self.add_row_btn.clicked.connect(self.add_row)
        self.del_row_btn.clicked.connect(self.delete_row)
        self.sort_btn.clicked.connect(self.sort_by_index)
        self.output_btn.clicked.connect(self.output_to_excel)
        self.save_btn.clicked.connect(self.save_progress)  # 新的保存进度按钮
        self.undo_btn.clicked.connect(self.undo_last_action)

        # 连接文件菜单动作
        open_excel_action.triggered.connect(self.open_excel)
        load_progress_action.triggered.connect(self.load_progress)

    def setup_keyboard_shortcuts(self):
        """设置Excel风格的键盘快捷键"""
        # 复制快捷键 (Ctrl+C)
        copy_action = QAction("复制", self)
        copy_action.setShortcut(QKeySequence.Copy)
        copy_action.triggered.connect(self.copy_selection)
        self.addAction(copy_action)

        # 粘贴快捷键 (Ctrl+V)
        paste_action = QAction("粘贴", self)
        paste_action.setShortcut(QKeySequence.Paste)
        paste_action.triggered.connect(self.paste_to_selection)
        self.addAction(paste_action)

        # 剪切快捷键 (Ctrl+X)
        cut_action = QAction("剪切", self)
        cut_action.setShortcut(QKeySequence.Cut)
        cut_action.triggered.connect(self.cut_selection)
        self.addAction(cut_action)

        # 保存快捷键 (Ctrl+S)
        save_action = QAction("保存", self)
        save_action.setShortcut(QKeySequence.Save)
        save_action.triggered.connect(self.save_progress)
        self.addAction(save_action)

        # 全选快捷键 (Ctrl+A)
        select_all_action = QAction("全选", self)
        select_all_action.setShortcut(QKeySequence.SelectAll)
        select_all_action.triggered.connect(self.select_all_cells)
        self.addAction(select_all_action)

        # 打开文件快捷键 (Ctrl+O)
        open_action = QAction("打开Excel", self)
        open_action.setShortcut(QKeySequence.Open)
        open_action.triggered.connect(self.open_excel)
        self.addAction(open_action)

        # 添加撤回功能的快捷键
        undo_action = QAction("撤回", self)
        undo_action.setShortcut(QKeySequence.Undo)  # Ctrl+Z
        undo_action.triggered.connect(self.undo_last_action)
        self.addAction(undo_action)

    def show_context_menu(self, position):
        """显示右键菜单，带撤回选项和清空单元格选项"""
        context_menu = QMenu(self)

        # 添加操作选项
        copy_action = context_menu.addAction("复制")
        paste_action = context_menu.addAction("粘贴")
        cut_action = context_menu.addAction("剪切")
        clear_cells_action = context_menu.addAction("清空单元格内容")  # 新增
        context_menu.addSeparator()

        add_row_action = context_menu.addAction("在此处添加行")
        delete_row_action = context_menu.addAction("删除选中行")
        context_menu.addSeparator()
        undo_action = context_menu.addAction("撤回上一步操作")

        # 连接信号
        copy_action.triggered.connect(self.copy_selection)
        paste_action.triggered.connect(self.paste_to_selection)
        cut_action.triggered.connect(self.cut_selection)
        clear_cells_action.triggered.connect(self.delete_selected_cells)  # 新增
        add_row_action.triggered.connect(self.add_row)
        delete_row_action.triggered.connect(self.delete_row)
        undo_action.triggered.connect(self.undo_last_action)

        # 禁用撤回选项如果没有历史
        undo_action.setEnabled(len(self.undo_stack) > 1)

        # 禁用清空单元格选项如果没有选中单元格
        clear_cells_action.setEnabled(len(self.table.selectedItems()) > 0)

        # 显示右键菜单
        context_menu.exec_(self.table.mapToGlobal(position))

    def copy_selection(self):
        """复制选中的单元格内容到剪贴板"""
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # 确定选中区域的范围
        rows = set()
        cols = set()
        for item in selected_items:
            rows.add(item.row())
            cols.add(item.column())
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)

        # 构建制表符分隔的文本
        text = ""
        for r in range(min_row, max_row + 1):
            row_texts = []
            for c in range(min_col, max_col + 1):
                item = self.table.item(r, c)
                if item and item in selected_items:
                    row_texts.append(item.text())
                else:
                    row_texts.append("")
            text += "\t".join(row_texts) + "\n"

        # 设置到剪贴板
        QApplication.clipboard().setText(text)

    def paste_to_selection(self):
        """从剪贴板粘贴到表格（优化撤回功能）"""
        # 获取剪贴板内容
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return

        # 获取当前选中单元格
        current_row = self.table.currentRow()
        current_col = self.table.currentColumn()
        if current_row < 0 or current_col < 0:
            return

        # 保存先前状态
        self.save_state()

        # 暂时断开信号连接
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 解析剪贴板数据
        rows = text.strip().split('\n')

        # 记录受影响的行数和单元格数
        affected_rows = []
        cell_count = 0

        # 粘贴数据
        for i, row_text in enumerate(rows):
            row_index = current_row + i
            affected_rows.append(row_index)

            # 如果需要添加新行
            while row_index >= self.table.rowCount():
                self.add_row()

            columns = row_text.split('\t')
            for j, cell_text in enumerate(columns):
                col_index = current_col + j

                # 确保列索引有效并且不是序号列
                if col_index < self.table.columnCount():
                    if col_index == 0:  # 保护序号列
                        continue
                    self.table.setItem(row_index, col_index, QTableWidgetItem(cell_text))
                    cell_count += 1

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 更新状态栏
        self.statusBar().showMessage(f"已粘贴 {cell_count} 个单元格内容到 {len(affected_rows)} 行")

    def cut_selection(self):
        """剪切选中内容（作为一个整体操作保存到撤回历史）"""
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # 保存当前状态以支持撤回
        self.save_state()

        # 复制到剪贴板
        self.copy_selection()

        # 暂时断开itemChanged信号，避免每个单元格修改都保存状态
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 清空选中的单元格（除了序号列）
        for item in selected_items:
            if item.column() != 0:  # 保护序号列
                item.setText("")

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        self.statusBar().showMessage(f"已剪切 {len(selected_items)} 个单元格内容")

    def delete_selected_cells(self):
        """删除选中单元格的内容（作为一个整体操作保存到撤回历史）"""
        # 获取选中的单元格
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # 保存当前状态以支持撤回
        self.save_state()

        # 暂时断开itemChanged信号，避免每个单元格修改都保存状态
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 清空选中的单元格内容（除序号列外）
        count = 0
        for item in selected_items:
            if item.column() != 0:  # 不删除序号列的内容
                item.setText("")
                count += 1

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 更新状态栏信息
        self.statusBar().showMessage(f"已清空 {count} 个单元格内容")

    def select_all_cells(self):
        """全选表格单元格"""
        self.table.selectAll()

    def keyPressEvent(self, event):
        """增强键盘导航处理，添加Delete键删除单元格内容功能"""
        # 处理Delete键删除单元格内容
        if event.key() == Qt.Key_Delete:
            if self.table.hasFocus() and self.table.selectedItems():
                self.delete_selected_cells()
                return

        # 处理剪切快捷键 (Ctrl+X)
        if event.key() == Qt.Key_X and event.modifiers() & Qt.ControlModifier:
            if self.table.hasFocus() and self.table.selectedItems():
                self.cut_selection()
                return

        # 处理回车键导航（编辑完成后移动到下一行）
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            current_row = self.table.currentRow()
            current_col = self.table.currentColumn()

            # 如果不是最后一行，移动到下一行
            if current_row < self.table.rowCount() - 1:
                self.table.setCurrentCell(current_row + 1, current_col)
            else:
                # 是最后一行，添加新行并移动
                self.add_row()
                self.table.setCurrentCell(current_row + 1, current_col)
            return

        # 处理Tab键导航
        if event.key() == Qt.Key_Tab:
            current_row = self.table.currentRow()
            current_col = self.table.currentColumn()

            # 移动到下一列或下一行
            if current_col < self.table.columnCount() - 1:
                self.table.setCurrentCell(current_row, current_col + 1)
            elif current_row < self.table.rowCount() - 1:
                self.table.setCurrentCell(current_row + 1, 0)
            return

        # 调用父类方法处理其他按键
        super().keyPressEvent(event)

    def open_excel(self):
        """打开Excel文件并导入到表格"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "打开Excel文件",
            "",
            "Excel Files (*.xlsx *.xls *.csv);;All Files (*)",
            options=options
        )

        self.save_state()

        if not file_path:
            return

        try:
            # 根据文件类型选择读取方式
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)

            # 清空当前表格
            self.table.setRowCount(0)

            # 获取列名
            headers = df.columns.tolist()

            # 检查是否有足够的列
            if len(headers) < 3:
                QMessageBox.warning(self, "警告", "Excel文件至少需要3列内容（序号、学院、财务金额）！")
                return

            # 添加数据到表格
            for idx, row in df.iterrows():
                table_row = self.table.rowCount()
                self.table.insertRow(table_row)

                # 序号
                seq_num = str(idx + 1)
                if len(headers) > 0 and not pd.isna(row[0]):
                    seq_num = str(row[0])
                self.table.setItem(table_row, 0, NumericTableWidgetItem(seq_num))

                # 学院名称
                school_name = ""
                if len(headers) > 1 and not pd.isna(row[1]):
                    school_name = str(row[1])
                self.table.setItem(table_row, 1, QTableWidgetItem(school_name))

                # 财务金额
                finance_amount = ""
                if len(headers) > 2 and not pd.isna(row[2]):
                    finance_amount = str(row[2])
                self.table.setItem(table_row, 2, QTableWidgetItem(finance_amount))

                # 是否补交 (如果有)
                supplement = ""
                if len(headers) > 3 and not pd.isna(row[3]):
                    supplement = str(row[3])
                self.table.setItem(table_row, 3, QTableWidgetItem(supplement))

            # 更新序号并排序
            self.update_row_numbers()
            self.sort_by_index()

            # 显示成功消息
            QMessageBox.information(self, "成功", f"已成功导入Excel文件：{os.path.basename(file_path)}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入Excel文件时发生错误：{str(e)}")

    def init_option_page(self):
        self.year_input = QLineEdit(self)
        self.month_input = QLineEdit(self)
        self.day_input = QLineEdit(self)

        self.year_input.setPlaceholderText("请输入团费年份（仅数字）")
        self.month_input.setPlaceholderText("请输入团费月份（仅数字）")
        self.day_input.setPlaceholderText("请输入落款日期（x年x月x日）")

        self.year_input.setStyleSheet("color: black; background-color: #e8f5e9; padding: 10px; font-size: 16px;")
        self.month_input.setStyleSheet("color: black; background-color: #e8f5e9; padding: 10px; font-size: 16px;")
        self.day_input.setStyleSheet("color: black; background-color: #e8f5e9; padding: 10px; font-size: 16px;")

        option_layout = QVBoxLayout(self.option_page)
        header_label = QLabel("选项设置", self)
        header_label.setStyleSheet("font-size: 28px; font-weight: bold; padding: 15px;")
        header_label.setAlignment(Qt.AlignCenter)
        option_layout.addWidget(header_label)

        input_layout = QHBoxLayout()
        input_layout.addWidget(self.year_input)
        input_layout.addWidget(self.month_input)
        input_layout.addWidget(self.day_input)
        option_layout.addLayout(input_layout)

    # 修改create_button方法支持图标
    def create_button(self, text, color, icon_file=None, base_path=None):
        """创建美观的按钮，支持图标"""
        btn = QPushButton(text)

        # 设置图标（如果存在）
        if icon_file and base_path:
            icon_path = os.path.join(base_path, icon_file)
            if os.path.exists(icon_path):
                btn.setIcon(QIcon(icon_path))
                btn.setIconSize(QSize(20, 20))

        # 应用现代按钮样式
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color}; 
                color: white; 
                padding: 10px; 
                font-size: 16px;
                border-radius: 4px;
                min-width: 120px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background-color: {color}ee;  /* 使用透明度创建悬停效果 */
            }}
            QPushButton:pressed {{
                background-color: {color}cc;  /* 使用透明度创建按下效果 */
            }}
        """)

        return btn

    # ============ 业务逻辑 ============

    def add_row(self):
        """添加新行（优化版，需要先选中单元格）"""
        # 确保有单元格被选中
        current_row = self.table.currentRow()
        if current_row == -1:
            self.statusBar().showMessage("请先选中单元格以确定添加位置")
            return

        # 保存当前状态
        self.save_state()

        # 暂时断开信号连接
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 执行添加行操作
        self.table.insertRow(current_row + 1)  # 在当前行之后添加

        # 设置新行单元格
        for col in range(self.table.columnCount()):
            if col == 0:
                self.table.setItem(current_row + 1, col, NumericTableWidgetItem(""))
            else:
                self.table.setItem(current_row + 1, col, QTableWidgetItem(""))

        # 更新所有行的序号
        self.update_row_numbers()

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 设置焦点到新行的第一个非序号列
        self.table.setCurrentCell(current_row + 1, 1)

        # 更新状态栏
        self.statusBar().showMessage(f"已添加新行（位置：{current_row + 2}）")

    def delete_row(self):
        """删除选中行（优化版）"""
        # 获取当前选中行
        current_row = self.table.currentRow()
        if current_row < 0:
            self.statusBar().showMessage("请先选中要删除的行")
            return

        # 保存当前状态
        self.save_state()

        # 暂时断开信号连接
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 记住删除前的有效行数和当前行
        row_count_before = self.table.rowCount()

        # 删除行
        self.table.removeRow(current_row)

        # 更新序号
        self.update_row_numbers()

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 设置焦点到合适的位置
        if self.table.rowCount() > 0:
            # 如果删除的是最后一行，选择新的最后一行
            if current_row >= self.table.rowCount():
                self.table.setCurrentCell(self.table.rowCount() - 1, 1)
            else:
                # 否则选择同一位置
                self.table.setCurrentCell(current_row, 1)

        # 更新状态栏
        self.statusBar().showMessage(f"已删除第 {current_row + 1} 行")

    def sort_by_index(self):
        self.table.sortItems(0, Qt.AscendingOrder)

    def output_to_excel(self):
        row_count = self.table.rowCount()
        data = []
        for row in range(row_count):
            row_data = []
            for col in range(3):
                item = self.table.item(row, col)
                cell_text = item.text() if item else ""
                row_data.append(cell_text)

            month_value = self.month_input.text()
            year_value = self.year_input.text()
            day_value = self.day_input.text()

            supplement_item = self.table.item(row, 3)
            if supplement_item and "补" in supplement_item.text():
                month_value = supplement_item.text()

            row_data.extend([month_value, year_value, day_value])
            data.append(row_data)

        base_path = os.path.dirname(os.path.abspath(__file__))
        save_dir = os.path.join(base_path, "../input")
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        file_path = os.path.join(save_dir, "data.xlsx")

        try:
            columns = ["序号", "学院", "财务金额", "团费月份", "团费年份", "落款日期"]
            df = pd.DataFrame(data, columns=columns)
            df.to_excel(file_path, index=False)

            workbook = load_workbook(file_path)
            worksheet = workbook.active
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            workbook.save(file_path)

            QMessageBox.information(self, "成功", "文件已成功保存！")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存文件时发生错误: {str(e)}")

    def save_progress(self):
        row_count = self.table.rowCount()
        data = []
        for row in range(row_count):
            row_data = []
            for col in range(4):
                item = self.table.item(row, col)
                cell_text = item.text() if item else ""
                row_data.append(cell_text)
            data.append(row_data)

        options_data = {
            "year": self.year_input.text(),
            "month": self.month_input.text(),
            "day": self.day_input.text()
        }

        save_data = {
            "table_data": data,
            "options_data": options_data
        }

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "保存进度", "", "JSON 文件 (*.json)", options=options)
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(save_data, f, ensure_ascii=False, indent=4)
                QMessageBox.information(self, "成功", "进度已成功保存！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"保存进度时发生错误: {str(e)}")

    def load_progress(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "导入进度", "", "JSON 文件 (*.json)", options=options)
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    load_data = json.load(f)

                self.table.setRowCount(0)
                for row_data in load_data["table_data"]:
                    current_row = self.table.rowCount()
                    self.table.insertRow(current_row)

                    numeric_item = NumericTableWidgetItem(row_data[0])
                    self.table.setItem(current_row, 0, numeric_item)
                    for col in range(1, 4):
                        self.table.setItem(current_row, col, QTableWidgetItem(row_data[col]))

                self.year_input.setText(load_data["options_data"]["year"])
                self.month_input.setText(load_data["options_data"]["month"])
                self.day_input.setText(load_data["options_data"]["day"])

                QMessageBox.information(self, "成功", "进度已成功导入！")
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导入进度时发生错误: {str(e)}")

    def update_buttons_state(self):
        """根据当前选择状态更新按钮的启用/禁用状态"""
        has_selection = self.table.currentRow() >= 0

        # 根据是否有选中的行来启用/禁用按钮
        self.add_row_btn.setEnabled(has_selection)
        self.del_row_btn.setEnabled(has_selection)

        # 撤回按钮根据历史记录数量启用/禁用
        self.undo_btn.setEnabled(len(self.undo_stack) > 1)

    def update_row_numbers(self):
        """更新所有行的序号（优化版）"""
        # 暂时断开信号连接
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 更新所有行的序号
        for row in range(self.table.rowCount()):
            item = NumericTableWidgetItem(str(row + 1))
            self.table.setItem(row, 0, item)

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

    def save_state(self):
        """保存当前表格状态到撤回历史"""
        # 获取当前表格数据
        table_data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                text = item.text() if item else ""
                row_data.append(text)
            table_data.append(row_data)

        # 限制撤回历史大小
        if len(self.undo_stack) >= self.max_undo_steps:
            self.undo_stack.pop(0)  # 移除最旧的状态

        # 添加当前状态到历史
        self.undo_stack.append(table_data)

        # 更新状态栏，并确保undo_btn已存在
        if len(self.undo_stack) > 1:
            self.statusBar().showMessage(f"已保存操作，可撤回（{len(self.undo_stack) - 1}步）")
            # 检查undo_btn是否存在
            if hasattr(self, 'undo_btn'):
                self.undo_btn.setEnabled(True)
        else:
            self.statusBar().showMessage("无法撤回")
            # 检查undo_btn是否存在
            if hasattr(self, 'undo_btn'):
                self.undo_btn.setEnabled(False)

    def undo_last_action(self):
        """改进的撤回功能，支持多步撤回和保持选中位置"""
        if len(self.undo_stack) <= 1:
            # 没有足够的历史可撤回
            self.statusBar().showMessage("没有可撤回的操作")
            self.undo_btn.setEnabled(False)
            return

        # 记住当前选中的单元格位置
        current_row = self.table.currentRow()
        current_col = self.table.currentColumn()

        # 设置标志，表示正在执行撤回操作
        self._is_undoing = True

        # 获取上一步的状态
        current_state = self.undo_stack.pop()  # 弹出当前状态
        previous_state = self.undo_stack[-1]  # 获取上一个状态（现在是最后一个）

        # 暂时断开表格信号以避免触发item_changed
        try:
            self.table.itemChanged.disconnect(self.on_item_changed)
        except TypeError:
            pass  # 如果信号未连接，忽略错误

        # 调整表格行数
        row_count_after = len(previous_state)
        row_count_before = self.table.rowCount()

        # 调整表格行数
        if row_count_before > row_count_after:
            # 需要删除行
            while self.table.rowCount() > row_count_after:
                self.table.removeRow(self.table.rowCount() - 1)
        elif row_count_before < row_count_after:
            # 需要添加行
            while self.table.rowCount() < row_count_after:
                self.table.insertRow(self.table.rowCount())

        # 恢复单元格数据
        for row in range(row_count_after):
            for col in range(min(len(previous_state[row]), self.table.columnCount())):
                text = previous_state[row][col]
                if col == 0:
                    item = NumericTableWidgetItem(text)
                else:
                    item = QTableWidgetItem(text)
                self.table.setItem(row, col, item)

        # 重新连接信号
        self.table.itemChanged.connect(self.on_item_changed)

        # 智能设置焦点位置
        if self.table.rowCount() > 0:
            # 如果原来选中的行仍然存在，保持该位置
            if current_row >= 0 and current_row < self.table.rowCount() and current_col >= 0 and current_col < self.table.columnCount():
                self.table.setCurrentCell(current_row, current_col)
            # 否则选择一个合理的位置
            elif self.table.rowCount() > 0:
                # 优先考虑靠近原位置的行
                self.table.setCurrentCell(min(current_row, self.table.rowCount() - 1),
                                          min(current_col if current_col >= 0 else 1, self.table.columnCount() - 1))

        # 清除正在执行撤回操作的标志
        self._is_undoing = False

        # 更新按钮状态
        self.update_buttons_state()

        # 更新状态栏
        remaining_steps = len(self.undo_stack) - 1
        if remaining_steps > 0:
            self.statusBar().showMessage(f"已撤回上一步操作，还可撤回{remaining_steps}步")
        else:
            self.statusBar().showMessage("已撤回到初始状态")

    def on_item_changed(self, item):
        # 动态添加属性
        if not hasattr(self, '_is_undoing'):
            self._is_undoing = False

        # 忽略序号列的变更
        if item.column() == 0:
            return

        # 如果正在执行撤回操作，不保存状态
        if self._is_undoing:
            return

        # 保存修改后的状态
        self.save_state()


if __name__ == "__main__":
    # 设置全局异常处理器来显示详细错误
    import traceback


    def exception_hook(exctype, value, tb):
        print(''.join(traceback.format_exception(exctype, value, tb)))
        sys.exit(1)


    sys.excepthook = exception_hook

    app = QApplication(sys.argv)

    # 暂时禁用qt_material主题，看是否是主题导致问题
    # apply_stylesheet(app, theme='light_blue.xml')

    try:
        window = ModernTableApp()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"启动失败: {e}")
        traceback.print_exc()
