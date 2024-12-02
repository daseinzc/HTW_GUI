import sys
import os
import logging
from openpyxl import load_workbook


def excelcatch():
    # 判断程序是否是通过打包后的 exe 启动
    if getattr(sys, 'frozen', False):  # 如果是打包后的可执行文件
        base_path = sys._MEIPASS  # 获取打包后的临时目录
    else:
        base_path = os.path.dirname(__file__)  # 开发时路径

    # 使用相对路径：从当前目录进入 'input' 文件夹
    data_path = os.path.join(base_path, '..', '..', '..', 'input', 'data.xlsx')
    logging.info(f"当前工作路径：{base_path}")
    logging.info(f"尝试读取文件路径：{data_path}")

    # 检查文件是否存在
    if not os.path.exists(data_path):
        logging.error(f"无法找到文件：{data_path}")
        return [], []  # 返回空的列表，或者可以抛出异常

    # 加载 Excel 文件
    try:
        wb = load_workbook(data_path)
    except Exception as e:
        logging.error(f"加载 Excel 文件失败: {e}")
        return [], []  # 返回空列表，或者可以抛出异常

    ws = wb.active
    ls_hang = []
    ls_school = []

    # 遍历 Excel 表中的所有行，跳过表头（从第二行开始）
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            logging.warning(f"发现不完整的数据行: {row}")
            continue  # 跳过不完整的行

        # 确保每一行都没有空值
        if any(cell is None for cell in row):
            logging.warning(f"发现空值的行: {row}")
            continue  # 跳过包含空值的行

        ls_hang.append(row)
        xuhao, school, money, month, year, due_day = row
        ls_school.append(school)

    return ls_hang, ls_school
